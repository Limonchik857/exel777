"""Чтение, проверка и экспорт таблиц (XLSX/CSV)."""

import csv
import re
import zipfile
from pathlib import Path

import pandas as pd

from django.conf import settings

from operations.validators import (
    ExportError,
    FileValidationError,
    TableReadError,
)

# Разрешённые расширения и MIME-типы
ALLOWED_EXTENSIONS = {".xlsx": "xlsx", ".csv": "csv"}
ALLOWED_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "text/plain",
    "application/csv",
    "application/x-csv",
}
# Браузеры часто шлют application/octet-stream; пропускаем только для
# разрешённых расширений — само по себе расширение уже проверено.
OCTET_STREAM = "application/octet-stream"

CSV_ENCODINGS = ["utf-8-sig", "cp1251", "utf-8", "latin-1"]

_SAFE_NAME_RE = re.compile(r"[^A-Za-zА-Яа-яЁё0-9_.\- ]+")

# Защита XLSX-контейнера от ZIP-бомб.
XLSX_MAX_ENTRIES = 10000
XLSX_MAX_UNCOMPRESSED_TOTAL = 2 * 1024**3  # 2 GiB суммарный распакованный размер
XLSX_MAX_COMPRESSION_RATIO = 500

# Символы, с которых начинаются формулы в Excel (CSV injection).
CSV_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def detect_file_type(filename):
    """Возвращает тип файла ('xlsx'/'csv') по расширению или None."""
    suffix = Path(filename).suffix.lower()
    return ALLOWED_EXTENSIONS.get(suffix)


def get_safe_original_name(filename):
    """Оставляет только безопасные символы в имени файла."""
    name = Path(filename).name
    name = _SAFE_NAME_RE.sub("_", name)
    return name.strip() or "file"


def validate_uploaded(request_file, max_size):
    """Проверяет расширение, MIME и размер. Возвращает (file_type, safe_name)."""
    original_name = request_file.name or ""
    file_type = detect_file_type(original_name)
    if file_type is None:
        raise FileValidationError("Формат файла не поддерживается. Разрешены XLSX и CSV.")

    mime = getattr(request_file, "content_type", "") or ""
    if mime and mime != OCTET_STREAM and mime not in ALLOWED_MIME:
        raise FileValidationError("Тип файла не распознан как таблица (XLSX или CSV).")

    if request_file.size > max_size:
        mb = max_size // (1024 * 1024)
        raise FileValidationError(f"Файл слишком большой. Максимальный размер — {mb} МБ.")

    if request_file.size == 0:
        raise FileValidationError("Файл пустой.")

    return file_type, get_safe_original_name(original_name)


def normalize_headers(df):
    """Приводит заголовки к единому виду: без пробелов по краям,
    без пустых имён и без дубликатов."""
    new_columns = []
    seen = {}
    for col in df.columns:
        name = str(col).strip()
        if not name or name.lower().startswith("unnamed:"):
            name = "unnamed"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        new_columns.append(name)
    df.columns = new_columns
    return df


def detect_encoding(path):
    sample = Path(path).read_bytes()[: 64 * 1024]
    for encoding in CSV_ENCODINGS:
        try:
            sample.decode(encoding)
            return encoding
        except (UnicodeDecodeError, Exception):
            continue
    return "utf-8"


def read_csv(path, nrows=None):
    encoding = detect_encoding(path)
    try:
        df = pd.read_csv(path, encoding=encoding, nrows=nrows)
    except Exception:
        raise TableReadError("Не удалось прочитать CSV-файл. Возможно, он повреждён.")
    return df


def validate_xlsx_security(path):
    """Проверяет ZIP-контейнер XLSX на признаки ZIP-бомбы.

    Отклоняет файлы с чрезмерным числом записей, слишком большим суммарным
    распакованным размером или аномальной степенью сжатия.
    """
    try:
        with zipfile.ZipFile(path, "r") as zf:
            infos = zf.infolist()
            if len(infos) > XLSX_MAX_ENTRIES:
                raise TableReadError(
                    "Файл отклонён: слишком много записей внутри XLSX-контейнера."
                )
            total_uncompressed = sum(info.file_size for info in infos)
            total_compressed = sum(info.compress_size for info in infos)
            if total_uncompressed > XLSX_MAX_UNCOMPRESSED_TOTAL:
                raise TableReadError(
                    "Файл отклонён: распакованный размер XLSX слишком велик."
                )
            if total_compressed > 0:
                ratio = total_uncompressed / total_compressed
                if ratio > XLSX_MAX_COMPRESSION_RATIO:
                    raise TableReadError(
                        "Файл отклонён: аномальная степень сжатия (подозрение на ZIP-бомбу)."
                    )
    except TableReadError:
        raise
    except zipfile.BadZipFile:
        raise TableReadError("Файл не является корректным XLSX (сломанный ZIP-контейнер).")
    except Exception:
        raise TableReadError("Не удалось проверить структуру XLSX-файла.")


def validate_table_size(df):
    """Проверяет таблицу на ограничения rows/columns/cells до тяжёлой обработки."""
    rows = len(df)
    cols = len(df.columns)
    if rows > settings.DATA_MAX_ROWS:
        raise TableReadError(
            f"Таблица слишком большая: {rows} строк. "
            f"Максимум — {settings.DATA_MAX_ROWS}."
        )
    if cols > settings.DATA_MAX_COLUMNS:
        raise TableReadError(
            f"Таблица слишком широкая: {cols} столбцов. "
            f"Максимум — {settings.DATA_MAX_COLUMNS}."
        )
    cells = rows * cols
    if cells > settings.DATA_MAX_CELLS:
        raise TableReadError(
            f"Таблица слишком большая: {cells} ячеек. "
            f"Максимум — {settings.DATA_MAX_CELLS}."
        )


def read_excel(path, nrows=None):
    validate_xlsx_security(path)
    try:
        df = pd.read_excel(path, engine="openpyxl", nrows=nrows)
    except Exception:
        raise TableReadError("Не удалось прочитать файл. Возможно, он повреждён или не является Excel.")
    return df


def read_table(path, file_type, nrows=None):
    """Читает таблицу в DataFrame с нормализованными заголовками."""
    if file_type == "csv":
        df = read_csv(path, nrows=nrows)
    elif file_type == "xlsx":
        df = read_excel(path, nrows=nrows)
    else:
        raise TableReadError("Неизвестный тип файла.")

    if df is None or df.empty:
        raise TableReadError("Файл пустой: в нём нет данных.")

    df = normalize_headers(df)
    df.columns = [str(c) for c in df.columns]
    validate_table_size(df)
    return df


def count_csv_rows(path):
    try:
        encoding = detect_encoding(path)
        with open(path, "r", encoding=encoding, newline="") as handle:
            return max(sum(1 for _ in csv.reader(handle)) - 1, 0)
    except Exception:
        return None


def count_xlsx_rows(path):
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            return max((ws.max_row or 1) - 1, 0)
        finally:
            wb.close()
    except Exception:
        return None


def table_info(path, file_type):
    """Возвращает (rows_count, columns, columns_count) без полного чтения."""
    if file_type == "csv":
        rows = count_csv_rows(path)
        try:
            sample = read_csv(path, nrows=0)
        except TableReadError:
            raise
        columns = [str(c).strip() for c in sample.columns]
    else:
        rows = count_xlsx_rows(path)
        sample = read_excel(path, nrows=0)
        columns = [str(c).strip() for c in sample.columns]
    if not columns or all(c in ("", "unnamed") for c in columns):
        raise TableReadError("Не удалось прочитать таблицу: отсутствуют заголовки столбцов.")
    return rows, columns, len(columns)


def _safe_csv_value(value):
    """Экранирует значения, с которых начинаются формулы Excel (=, +, -, @)."""
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    if text.startswith(CSV_DANGEROUS_PREFIXES):
        return "'" + text
    return text


def export_dataframe(df, file_type, safe_csv=None):
    """Экспортирует DataFrame в байты (xlsx или csv). Возвращает (bytes, mime)."""
    try:
        if file_type == "xlsx":
            from io import BytesIO

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Data")
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return buffer.getvalue(), mime
        if file_type == "csv":
            from io import StringIO

            safe = settings.SAFE_CSV_EXPORT if safe_csv is None else safe_csv
            export_df = df
            if safe:
                export_df = df.copy()
                for col in export_df.columns:
                    export_df[col] = export_df[col].map(_safe_csv_value)
            buffer = StringIO()
            export_df.to_csv(buffer, index=False, encoding="utf-8")
            # BOM для корректного открытия в Excel
            data = buffer.getvalue().encode("utf-8-sig")
            return data, "text/csv; charset=utf-8"
    except Exception:
        raise ExportError("Произошла ошибка при формировании файла результата.")
    raise ExportError("Неизвестный формат экспорта.")


def suggest_result_name(source_name, file_type):
    stem = Path(source_name).stem
    suffix = ".xlsx" if file_type == "xlsx" else ".csv"
    return f"{stem}_cleaned{suffix}"


def cleanup_stale_processing(user_id, session_token):
    """Бест-эфорт очистка временных файлов старой сессии обработки."""
    root = processing_root(user_id, session_token)
    try:
        import shutil

        if root.exists():
            shutil.rmtree(root, ignore_errors=True)
    except Exception:
        pass


def processing_root(user_id, session_token):
    return settings.MEDIA_ROOT / "processing" / str(user_id) / session_token